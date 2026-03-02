#!/usr/bin/env python3
"""
Create a color-filled Excel workbook from existing snowfall/temperature CSV tables.

Rules (fill color):
- Snowfall: 1-14 cm -> light blue, 15+ cm -> light orange
- Temperature: <0 C -> light blue, 1-4 C -> light purple, 4+ C -> light red
"""

from __future__ import annotations

import argparse
import csv
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SNOW_ORANGE = PatternFill(fill_type="solid", fgColor="FFE7CC")
TEMP_BLUE = PatternFill(fill_type="solid", fgColor="CFE8FF")
TEMP_PURPLE = PatternFill(fill_type="solid", fgColor="E8D9FF")
TEMP_RED = PatternFill(fill_type="solid", fgColor="FFD6D6")


def to_float(value: str) -> Optional[float]:
    v = (value or "").strip()
    if v == "":
        return None
    try:
        return float(v)
    except ValueError:
        return None


def _interp_channel(a: int, b: int, t: float) -> int:
    return int(round(a + (b - a) * t))


def snowfall_fill(value: float) -> PatternFill:
    # 0-15: white -> light blue gradient, >15: light orange
    if value > 15:
        return SNOW_ORANGE
    v = min(max(value, 0.0), 15.0)
    t = v / 15.0
    # White (255,255,255) to light blue (207,232,255)
    r = _interp_channel(255, 207, t)
    g = _interp_channel(255, 232, t)
    b = _interp_channel(255, 255, t)
    rgb = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(fill_type="solid", fgColor=rgb)


def read_csv(path: str) -> List[Dict[str, str]]:
    with open(path, "r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def autosize_columns(ws) -> None:
    for col in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 10), 60)


def write_snowfall_sheet(wb: Workbook, rows: List[Dict[str, str]]) -> None:
    ws = wb.active
    ws.title = "snowfall"
    if not rows:
        ws.append(["No data"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    snow_cols = [h for h in headers if h.startswith("day_") and h.endswith("_cm")]
    if "week1_total_cm" in headers:
        snow_cols.append("week1_total_cm")
    if "week2_total_cm" in headers:
        snow_cols.append("week2_total_cm")
    snow_col_idx = {h: headers.index(h) + 1 for h in snow_cols}

    for r in rows:
        ws.append([r.get(h, "") for h in headers])
        row_idx = ws.max_row
        for h, col_idx in snow_col_idx.items():
            val = to_float(r.get(h, ""))
            if val is None:
                continue
            ws.cell(row=row_idx, column=col_idx).fill = snowfall_fill(val)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)


def write_temperature_sheet(wb: Workbook, rows: List[Dict[str, str]]) -> None:
    ws = wb.create_sheet("temperature")
    if not rows:
        ws.append(["No data"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    temp_cols = [h for h in headers if h.endswith("_max_c") or h.endswith("_min_c")]
    temp_col_idx = {h: headers.index(h) + 1 for h in temp_cols}

    for r in rows:
        ws.append([r.get(h, "") for h in headers])
        row_idx = ws.max_row
        for h, col_idx in temp_col_idx.items():
            val = to_float(r.get(h, ""))
            if val is None:
                continue
            if val < 0:
                ws.cell(row=row_idx, column=col_idx).fill = TEMP_BLUE
            elif 1 <= val <= 4:
                ws.cell(row=row_idx, column=col_idx).fill = TEMP_PURPLE
            elif val > 4:
                ws.cell(row=row_idx, column=col_idx).fill = TEMP_RED

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Apply fill colors to snowfall/temperature tables.")
    p.add_argument("--snowfall-csv", default="resorts_snowfall_daily.csv")
    p.add_argument("--temperature-csv", default="resorts_temperature_daily.csv")
    p.add_argument("--output-xlsx", default="resorts_colored.xlsx")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    snow_rows = read_csv(args.snowfall_csv)
    temp_rows = read_csv(args.temperature_csv)

    wb = Workbook()
    write_snowfall_sheet(wb, snow_rows)
    write_temperature_sheet(wb, temp_rows)
    wb.save(args.output_xlsx)
    print(f"Done: {args.output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
